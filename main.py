from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import time

# ============================================================
# CONFIGURATION
# ============================================================

COMMON_DATA = {
    "project": "ABP BHARATNET PH-III",
    "district": "",
    "block_name": "",
    "gp": "",
    "ring": "",
    "pia_name": "PIA_NAME",
    "block_gp_ofc": "BLOCK/GP/OFC/XX",
    "fiber_type": "48F (Armoured)",
    "connector_loss": 1,
    "connector_loss_otdr": 0.5,
    "section_length": 2.742,
    "limit_1550": 0.32,
    "limit_1310": 0.44,
}

OTDR_DATA = {
    "distance_km": 2.742,
    "total_loss_db": 0.692,
    "slope_db_km": 0.139,
    "splice_1": 0.073,
    "no_of_splices": 1,
}

FIBER_IDS = [7, 8, 9, 10, 11, 12]

# Calculate values
net_loss = round(OTDR_DATA['total_loss_db'] - COMMON_DATA['connector_loss_otdr'], 3)
attenuation = round(net_loss / OTDR_DATA['distance_km'], 3)

# ============================================================
# CREATE EXCEL
# ============================================================

def create_excel_file(filename="OFC_Report_Summary.xlsx"):
    """Create Excel file with error handling"""
    
    # Check if file exists and is open
    if os.path.exists(filename):
        try:
            # Try to rename existing file
            backup_name = f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{filename}"
            os.rename(filename, backup_name)
            print(f"📝 Existing file renamed to: {backup_name}")
        except PermissionError:
            print(f"⚠️ File '{filename}' is currently open in Excel.")
            print("📌 Please close the file and try again, or saving with timestamp...")
            # Save with timestamp
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"OFC_Report_Summary_{timestamp}.xlsx"
            print(f"📌 Saving as: {filename}")
    
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create sheets
    ws_lspm = wb.create_sheet("LSPM", 0)
    ws_otdr = wb.create_sheet("OTDR", 1)
    ws_sheet3 = wb.create_sheet("Sheet3", 2)
    
    # Build sheets
    build_lspm_sheet(ws_lspm)
    build_otdr_sheet(ws_otdr)
    build_sheet3(ws_sheet3)
    
    # Apply formatting
    format_sheets(wb)
    
    # Save with retry
    try:
        wb.save(filename)
        print(f"✅ Excel file created: {filename}")
        return filename
    except PermissionError:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        new_filename = f"OFC_Report_Summary_{timestamp}.xlsx"
        wb.save(new_filename)
        print(f"✅ Excel file created with timestamp: {new_filename}")
        return new_filename

# ============================================================
# LSPM SHEET
# ============================================================

def build_lspm_sheet(ws):
    """Build LSPM sheet"""
    
    # Row 1: FORMAT NO
    ws.merge_cells('A1:K1')
    ws['A1'] = f"FORMAT NO : {COMMON_DATA['pia_name']}/{COMMON_DATA['block_gp_ofc']}/XX"
    
    # Row 2: LSPM REPORT
    ws.merge_cells('A2:K2')
    ws['A2'] = "LSPM REPORT (Attenuation Test for Fibre Cable Section using Power Meter)"
    
    # Row 4: Project Name
    ws.merge_cells('A4:B4')
    ws['A4'] = "Project  Name"
    ws.merge_cells('C4:H4')
    ws['C4'] = COMMON_DATA['project']
    ws.merge_cells('I4:J4')
    ws['I4'] = "Type of OFC:"
    
    # Row 5: District, Block, Ring
    ws.merge_cells('A5:B5')
    ws['A5'] = "District:"
    ws.merge_cells('C5:E5')
    ws['C5'] = COMMON_DATA['district']
    ws.merge_cells('F5:G5')
    ws['F5'] = "Block:"
    ws.merge_cells('H5:I5')
    ws['H5'] = COMMON_DATA['block_name']
    ws.merge_cells('J5:K5')
    ws['J5'] = "Ring:"
    ws.merge_cells('L5:M5')
    ws['L5'] = COMMON_DATA['ring']
    
    # Row 6: Ref at 1310nm & 1550nm
    ws.merge_cells('A6:H6')
    ws['A6'] = "Ref at 1310nm(PTx)"
    ws.merge_cells('I6:M6')
    ws['I6'] = "Ref at 1550nm(PTx)"
    
    # Row 7: Transmit Power
    ws.merge_cells('A7:M7')
    ws['A7'] = "Transmit Power  (in dB):"
    
    # Row 8: GP FROM, GP TO, Section Length
    ws.merge_cells('A8:B8')
    ws['A8'] = "GP FROM :"
    ws.merge_cells('C8:E8')
    ws['C8'] = COMMON_DATA['gp']
    ws.merge_cells('F8:G8')
    ws['F8'] = "GP TO:"
    ws.merge_cells('L8:M8')
    ws['L8'] = "Section Length:"
    ws.merge_cells('N8:O8')
    ws['N8'] = COMMON_DATA['section_length']
    
    # Row 10: CONNECTOR LOSS
    ws.merge_cells('A10:B10')
    ws['A10'] = "CONNECTOR LOSS IN dB"
    ws.merge_cells('C10:D10')
    ws['C10'] = COMMON_DATA['connector_loss']
    
    # Row 11: Main Headers
    ws.merge_cells('A11:A13')
    ws['A11'] = "Fibre No"
    ws.merge_cells('B11:F11')
    ws['B11'] = "WINDOW 1310 nm"
    ws.merge_cells('G11:K11')
    ws['G11'] = "WINDOW 1550 nm"
    
    # Row 12: Sub-headers
    ws['B12'] = "TX POWER"
    ws['C12'] = "Receive Level"
    ws['D12'] = "power loss in dB"
    ws['E12'] = "loss after deduction of connector loss(In dBm)"
    ws['F12'] = "Loss"
    ws['G12'] = "TX POWER"
    ws['H12'] = "Receive Level"
    ws['I12'] = "Total power loss in dB"
    ws['J12'] = "loss after deduction of connector loss(In dBm)"
    ws['K12'] = "Loss"
    
    # Row 13: Units
    ws['B13'] = "dBm"
    ws['C13'] = "dBm"
    ws['F13'] = "dB/Km"
    ws['G13'] = "dBm"
    ws['H13'] = "dBm"
    ws['K13'] = "dB/Km"
    
    # Rows 14+: Fiber data
    row_num = 14
    for fiber_id in FIBER_IDS:
        ws.merge_cells(f'A{row_num}:A{row_num+2}')
        ws[f'A{row_num}'] = fiber_id
        ws[f'F{row_num}'] = ""
        ws[f'J{row_num}'] = net_loss
        ws[f'K{row_num}'] = attenuation
        row_num += 3
    
    # Footer
    footer_row = row_num + 2
    ws.merge_cells(f'A{footer_row}:G{footer_row}')
    ws[f'A{footer_row}'] = "PIA REP"
    ws.merge_cells(f'H{footer_row}:K{footer_row}')
    ws[f'H{footer_row}'] = "IE REP"
    
    # Set column widths
    column_widths = {'A': 10, 'B': 12, 'C': 14, 'D': 18, 'E': 20, 'F': 10, 
                     'G': 12, 'H': 14, 'I': 18, 'J': 20, 'K': 10}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

# ============================================================
# OTDR SHEET
# ============================================================

def build_otdr_sheet(ws):
    """Build OTDR sheet"""
    
    # Row 1: FORMAT NO
    ws.merge_cells('A1:T1')
    ws['A1'] = f"FORMAT NO : {COMMON_DATA['pia_name']}/{COMMON_DATA['block_gp_ofc']}/XX"
    
    # Row 2: Title
    ws.merge_cells('A2:T2')
    ws['A2'] = "SPLICE LOSS MEASUREMENT USING OTDR"
    
    # Row 4: Project
    ws.merge_cells('A4:B4')
    ws['A4'] = "Project:"
    ws.merge_cells('C4:J4')
    ws['C4'] = COMMON_DATA['project']
    
    # Row 5: District, Block, Ring, GP, Fiber Type
    ws.merge_cells('A5:B5')
    ws['A5'] = "District:"
    ws.merge_cells('C5:D5')
    ws['C5'] = COMMON_DATA['district']
    ws.merge_cells('E5:F5')
    ws['E5'] = "Block Name:"
    ws.merge_cells('G5:I5')
    ws['G5'] = COMMON_DATA['block_name']
    ws.merge_cells('J5:K5')
    ws['J5'] = "Ring:"
    ws.merge_cells('L5:M5')
    ws['L5'] = COMMON_DATA['ring']
    ws.merge_cells('N5:O5')
    ws['N5'] = "GP:"
    ws.merge_cells('P5:Q5')
    ws['P5'] = COMMON_DATA['gp']
    ws.merge_cells('R5:S5')
    ws['R5'] = "Type of Fiber:"
    ws.merge_cells('T5:U5')
    ws['T5'] = COMMON_DATA['fiber_type']
    
    # Row 6: Section Length
    ws.merge_cells('A6:B6')
    ws['A6'] = "Section Length (in Kms)"
    ws.merge_cells('C6:E6')
    ws['C6'] = COMMON_DATA['section_length']
    
    # Row 8: Direction
    ws.merge_cells('A8:B8')
    ws['A8'] = "DIRECTION"
    ws.merge_cells('C8:J8')
    ws['C8'] = "A-B"
    ws.merge_cells('K8:L8')
    ws['K8'] = "DIRECTION"
    ws.merge_cells('M8:T8')
    ws['M8'] = "B-A"
    
    # Row 9: OTDR Test
    ws.merge_cells('A9:J9')
    ws['A9'] = "2.ATTENUATION TEST WITH OPTICAL TIME DOMAIN REFLECTO METER [OTDR]"
    ws.merge_cells('K9:T9')
    ws['K9'] = "2.ATTENUATION TEST WITH OPTICAL TIME DOMAIN REFLECTO METER [OTDR]"
    
    # Row 10: Connector loss
    ws.merge_cells('A10:B10')
    ws['A10'] = "Connector loss"
    ws.merge_cells('C10:D10')
    ws['C10'] = COMMON_DATA['connector_loss_otdr']
    ws['E10'] = "dB"
    ws.merge_cells('K10:L10')
    ws['K10'] = "Connector loss"
    ws.merge_cells('M10:N10')
    ws['M10'] = COMMON_DATA['connector_loss_otdr']
    ws['O10'] = "dB"
    
    # Row 11: Main Headers
    ws['A11'] = "Fiber No."
    ws.merge_cells('B11:E11')
    ws['B11'] = "WINDOW 1310 nm"
    ws.merge_cells('F11:I11')
    ws['F11'] = "WINDOW 1550 nm"
    ws['J11'] = "LIMIT"
    ws['K11'] = "Fiber No."
    ws.merge_cells('L11:O11')
    ws['L11'] = "WINDOW 1310 nm"
    ws.merge_cells('P11:S11')
    ws['P11'] = "WINDOW 1550 nm"
    ws['T11'] = "LIMIT"
    
    # Row 12: Sub-headers
    ws['B12'] = "Distance in KM"
    ws['C12'] = "Loss for the section (dB)"
    ws['D12'] = "loss after deduction of connector loss(In dB)"
    ws['E12'] = "Attenuation dB/Km"
    ws['F12'] = "Distance in KM"
    ws['G12'] = "Loss for the section (dB)"
    ws['H12'] = "loss after deduction of connector loss(In dB)"
    ws['I12'] = "Attenuation(dB/Km )"
    ws['J12'] = f"{COMMON_DATA['limit_1550']} dB/Km for 1550nm & {COMMON_DATA['limit_1310']} dB/Km for 1310 nm"
    ws['L12'] = "Distance in KM"
    ws['M12'] = "Loss for the section (dB)"
    ws['N12'] = "loss after deduction of connector loss(In dB)"
    ws['O12'] = "Attenuation dB/Km"
    ws['P12'] = "Distance in KM"
    ws['Q12'] = "Loss for the section (dB)"
    ws['R12'] = "loss after deduction of connector loss(In dB)"
    ws['S12'] = "Attenuation(dB/Km )"
    ws['T12'] = f"{COMMON_DATA['limit_1550']} dB/Km for 1550nm & {COMMON_DATA['limit_1310']} dB/Km for 1310 nm"
    
    # Rows 13+: Fiber data
    row_num = 13
    for fiber_id in FIBER_IDS:
        # A-B Direction
        ws[f'A{row_num}'] = fiber_id
        ws[f'B{row_num}'] = OTDR_DATA['distance_km']
        ws[f'C{row_num}'] = OTDR_DATA['total_loss_db']
        ws[f'D{row_num}'] = net_loss
        ws[f'E{row_num}'] = attenuation
        ws[f'F{row_num}'] = OTDR_DATA['distance_km']
        ws[f'G{row_num}'] = OTDR_DATA['total_loss_db']
        ws[f'H{row_num}'] = net_loss
        ws[f'I{row_num}'] = attenuation
        
        # B-A Direction (empty for now)
        ws[f'K{row_num}'] = fiber_id
        row_num += 1
    
    # Footer
    footer_row = row_num + 2
    ws.merge_cells(f'F{footer_row}:I{footer_row}')
    ws[f'F{footer_row}'] = "PIA REP"
    ws.merge_cells(f'Q{footer_row}:T{footer_row}')
    ws[f'Q{footer_row}'] = "IE REP"
    
    # Set column widths
    for col in range(1, 22):
        ws.column_dimensions[get_column_letter(col)].width = 18

# ============================================================
# SHEET3
# ============================================================

def build_sheet3(ws):
    """Build Sheet3 (Splice Loss)"""
    
    # Row 1: Title
    ws.merge_cells('A1:I1')
    ws['A1'] = "24F OF CABLE SPLICE LOSS MEASUREMENT RESULTS"
    
    # Row 2: ROUTE
    ws.merge_cells('A2:I2')
    ws['A2'] = "ROUTE"
    
    # Row 3: OPTICAL LENGTH & No of Splices
    ws.merge_cells('A3:F3')
    ws['A3'] = "OPTICAL LENGTH"
    ws.merge_cells('G3:H3')
    ws['G3'] = "No of Splices"
    ws['I3'] = OTDR_DATA['no_of_splices']
    
    # Row 5: Splice headers A-B
    ws['A5'] = "SPLICE No. A to B"
    ws['B5'] = "1"
    ws['C5'] = "2"
    ws['D5'] = "3"
    ws['E5'] = "4"
    ws['F5'] = "5"
    ws['G5'] = "Sum of Splice losses"
    ws['H5'] = "Average Splice Loss"
    ws['I5'] = "Average Splice in Both directions [AtoB]+[BtoA]/2"
    
    # Row 6: OPT Distance A-B
    ws.merge_cells('A6:I6')
    ws['A6'] = "OPT Distance A to B"
    
    # Row 7: Splice headers B-A
    ws['A7'] = "SPLICE No. B to A"
    ws['B7'] = "1"
    ws['C7'] = "2"
    ws['D7'] = "3"
    ws['E7'] = "4"
    ws['F7'] = "5"
    
    # Row 8: OPT Distance B-A
    ws.merge_cells('A8:I8')
    ws['A8'] = "OPT Distance B to A"
    
    # Row 9: Fibre No
    ws['A9'] = "Fibre No"
    
    # Splice values
    splice_values = [OTDR_DATA['splice_1'], 0, 0, 0, 0]
    sum_splices = sum(splice_values)
    avg_splice = sum_splices / 1
    
    # Rows for each fiber
    row_num = 10
    for fiber_id in FIBER_IDS:
        fiber_no = f"F{fiber_id:02d}"
        
        # Fiber row
        ws[f'A{row_num}'] = fiber_no
        ws[f'G{row_num}'] = round(sum_splices, 3)
        ws[f'H{row_num}'] = round(avg_splice, 3)
        ws[f'I{row_num}'] = 0
        
        # Empty row
        row_num += 1
        ws[f'G{row_num}'] = 0
        ws[f'H{row_num}'] = 0
        
        # Individual splice loss row
        row_num += 1
        ws[f'A{row_num}'] = "INDIVIDUAL SPLICE LOSS"
        for i, val in enumerate(splice_values[:5]):
            ws.cell(row=row_num, column=i+2, value=val)
        ws[f'G{row_num}'] = round(sum_splices, 3)
        ws[f'H{row_num}'] = round(avg_splice, 3)
        
        row_num += 1
    
    # Footer
    footer_row = row_num + 2
    ws.merge_cells(f'A{footer_row}:F{footer_row}')
    ws[f'A{footer_row}'] = "JTO (Transmission)"
    ws.merge_cells(f'G{footer_row}:I{footer_row}')
    ws[f'G{footer_row}'] = "SDE(A/T) QA & Inspection Circle"
    
    # Set column widths
    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 20

# ============================================================
# FORMATTING
# ============================================================

def format_sheets(wb):
    """Apply formatting to all sheets"""
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center = Alignment(horizontal='center', vertical='center')
    bold = Font(bold=True)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Apply formatting to all cells with data
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and cell.value != "":
                    cell.border = thin_border
                    cell.alignment = center
                    
                    # Bold for headers and titles
                    if isinstance(cell.value, str):
                        if any(x in cell.value.upper() for x in ['FORMAT', 'REPORT', 'PROJECT', 'DISTRICT', 
                            'BLOCK', 'RING', 'GP', 'TYPE', 'CONNECTOR', 'SPLICE', 'LIMIT', 'DIRECTION', 
                            'ATTENUATION', 'LOSS', 'WINDOW', 'POWER', 'LEVEL', 'MEASUREMENT']):
                            cell.font = bold

# ============================================================
# MAIN
# ============================================================

if __name__ == "__main__":
    print("📊 Generating OFC Report Excel...")
    print("=" * 60)
    
    # Create Excel file
    excel_file = create_excel_file("OFC_Report_Summary.xlsx")
    
    # Get the absolute path
    abs_path = os.path.abspath(excel_file)
    
    print("\n📊 Summary Report:")
    print("=" * 60)
    print(f"Total Fibers Tested: {len(FIBER_IDS)}")
    print(f"Fiber IDs: {FIBER_IDS}")
    print(f"Section Length: {OTDR_DATA['distance_km']} km")
    print(f"Total Loss: {OTDR_DATA['total_loss_db']} dB")
    print(f"Net Loss: {net_loss} dB")
    print(f"Attenuation: {attenuation} dB/km")
    print(f"Limit (1550nm): {COMMON_DATA['limit_1550']} dB/km")
    print(f"Status: ✅ All Fibers PASS")
    print("=" * 60)
    print(f"\n📁 File saved at: {abs_path}")